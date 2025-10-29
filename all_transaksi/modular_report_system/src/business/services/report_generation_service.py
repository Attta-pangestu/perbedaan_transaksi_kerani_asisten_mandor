"""
Report Generation Service
Service for generating reports based on templates and data
"""

import logging
import threading
import time
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Callable
from pathlib import Path
import json

from business.models.report_template import ReportTemplate
from business.models.report_request import ReportRequest
from business.models.report_result import ReportResult
from infrastructure.exceptions.custom_exceptions import *


class ReportGenerationService:
    """
    Service for generating reports
    """
    
    def __init__(self):
        """Initialize report generation service"""
        self.logger = logging.getLogger(__name__)
        
        # Generation state
        self.is_generating = False
        self.current_request: Optional[ReportRequest] = None
        self.generation_thread: Optional[threading.Thread] = None
        self.cancel_requested = False
        
        # Callbacks
        self.progress_callback: Optional[Callable] = None
        self.completion_callback: Optional[Callable] = None
        self.error_callback: Optional[Callable] = None
        
        # Statistics
        self.generation_stats = {
            'total_generated': 0,
            'successful_generations': 0,
            'failed_generations': 0,
            'last_generation_time': None
        }
    
    def generate_report_async(self, request: ReportRequest,
                            progress_callback: Optional[Callable] = None,
                            completion_callback: Optional[Callable] = None,
                            error_callback: Optional[Callable] = None):
        """
        Generate report asynchronously
        
        Args:
            request: Report generation request
            progress_callback: Progress update callback
            completion_callback: Completion callback
            error_callback: Error callback
        """
        if self.is_generating:
            raise ReportGenerationError("Report generation already in progress")
        
        # Set callbacks
        self.progress_callback = progress_callback
        self.completion_callback = completion_callback
        self.error_callback = error_callback
        
        # Set state
        self.is_generating = True
        self.current_request = request
        self.cancel_requested = False
        
        # Start generation thread
        self.generation_thread = threading.Thread(
            target=self._generate_report_thread,
            args=(request,),
            daemon=True
        )
        self.generation_thread.start()
        
        self.logger.info(f"Started async report generation for template: {request.template.name}")
    
    def _generate_report_thread(self, request: ReportRequest):
        """
        Report generation thread
        
        Args:
            request: Report generation request
        """
        try:
            result = self._generate_report_sync(request)
            
            if not self.cancel_requested:
                self.generation_stats['successful_generations'] += 1
                self.generation_stats['last_generation_time'] = datetime.now()
                
                if self.completion_callback:
                    self.completion_callback(result)
            
        except Exception as e:
            self.generation_stats['failed_generations'] += 1
            error_message = str(e)
            
            self.logger.error(f"Report generation failed: {error_message}")
            
            if not self.cancel_requested and self.error_callback:
                self.error_callback(error_message)
        
        finally:
            self.generation_stats['total_generated'] += 1
            self.is_generating = False
            self.current_request = None
            self.generation_thread = None
    
    def _generate_report_sync(self, request: ReportRequest) -> ReportResult:
        """
        Generate report synchronously
        
        Args:
            request: Report generation request
            
        Returns:
            Report generation result
        """
        self.logger.info(f"Starting report generation for template: {request.template.name}")
        
        # Validate request
        self._validate_request(request)
        
        # Initialize result
        result = ReportResult(
            request=request,
            start_time=datetime.now(),
            status='in_progress'
        )
        
        try:
            # Step 1: Prepare data sources
            self._update_progress(1, "Mempersiapkan sumber data...", "Menginisialisasi koneksi database")
            self._prepare_data_sources(request)
            self._check_cancellation()
            
            # Step 2: Load and validate template
            self._update_progress(2, "Memuat template...", f"Memuat template: {request.template.name}")
            template_data = self._load_template_data(request.template)
            self._check_cancellation()
            
            # Step 3: Extract data
            self._update_progress(3, "Mengekstrak data...", "Menjalankan query database")
            raw_data = self._extract_data(request, template_data)
            self._check_cancellation()
            
            # Step 4: Process data
            self._update_progress(4, "Memproses data...", "Menerapkan logika bisnis")
            processed_data = self._process_data(raw_data, request, template_data)
            self._check_cancellation()
            
            # Step 5: Generate reports
            self._update_progress(5, "Membuat laporan...", "Menghasilkan file output")
            output_files = self._generate_output_files(processed_data, request, template_data)
            self._check_cancellation()
            
            # Step 6: Finalize
            self._update_progress(6, "Menyelesaikan...", "Menyimpan hasil dan membersihkan")
            self._finalize_generation(result, output_files)
            
            result.end_time = datetime.now()
            result.status = 'completed'
            result.output_files = output_files
            result.success = True
            
            self.logger.info(f"Report generation completed successfully")
            return result
            
        except Exception as e:
            result.end_time = datetime.now()
            result.status = 'failed'
            result.error_message = str(e)
            result.success = False
            raise
    
    def _validate_request(self, request: ReportRequest):
        """
        Validate report request
        
        Args:
            request: Report generation request
        """
        if not request.template:
            raise ValidationError("Template is required")
        
        if not request.template.is_valid():
            raise ValidationError("Template is not valid")
        
        if not request.selected_estates:
            raise ValidationError("At least one estate must be selected")
        
        if not request.selected_month:
            raise ValidationError("Month selection is required")
        
        # Validate month format
        try:
            datetime.strptime(request.selected_month, "%Y-%m")
        except ValueError:
            raise ValidationError("Invalid month format. Expected YYYY-MM")
    
    def _prepare_data_sources(self, request: ReportRequest):
        """
        Prepare data sources for report generation
        
        Args:
            request: Report generation request
        """
        # Simulate data source preparation
        time.sleep(0.5)
        
        # Here you would typically:
        # - Initialize database connections
        # - Validate estate configurations
        # - Check data availability
        
        self.logger.debug("Data sources prepared")
    
    def _load_template_data(self, template: ReportTemplate) -> Dict[str, Any]:
        """
        Load template data and configuration
        
        Args:
            template: Report template
            
        Returns:
            Template data dictionary
        """
        # Simulate template loading
        time.sleep(0.3)
        
        # Load template configuration
        template_data = {
            'template': template,
            'config': template.config,
            'ui_flow': template.ui_flow,
            'data_processing': template.data_processing,
            'report_format': template.report_format
        }
        
        self.logger.debug(f"Template data loaded for: {template.name}")
        return template_data
    
    def _extract_data(self, request: ReportRequest, template_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract data from data sources
        
        Args:
            request: Report generation request
            template_data: Template configuration data
            
        Returns:
            Raw data dictionary
        """
        # Simulate data extraction
        time.sleep(1.0)
        
        # Here you would typically:
        # - Execute database queries
        # - Filter data by estate and month
        # - Apply template-specific data extraction logic
        
        # Mock data for demonstration
        raw_data = {
            'estates': request.selected_estates,
            'month': request.selected_month,
            'transactions': self._generate_mock_transactions(request),
            'metadata': {
                'extraction_time': datetime.now(),
                'record_count': 100,  # Mock count
                'estates_processed': len(request.selected_estates)
            }
        }
        
        self.logger.debug(f"Data extracted for {len(request.selected_estates)} estates")
        return raw_data
    
    def _process_data(self, raw_data: Dict[str, Any], 
                     request: ReportRequest, 
                     template_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process raw data according to template logic
        
        Args:
            raw_data: Raw extracted data
            request: Report generation request
            template_data: Template configuration data
            
        Returns:
            Processed data dictionary
        """
        # Simulate data processing
        time.sleep(0.8)
        
        # Here you would typically:
        # - Apply business logic from template
        # - Perform calculations and aggregations
        # - Detect duplicates or anomalies
        # - Format data for reporting
        
        processed_data = {
            'summary': {
                'total_transactions': raw_data['metadata']['record_count'],
                'estates_count': len(raw_data['estates']),
                'month': raw_data['month'],
                'processing_time': datetime.now()
            },
            'details': raw_data['transactions'],
            'analysis': {
                'duplicates_found': 5,  # Mock analysis
                'anomalies_detected': 2,
                'data_quality_score': 95.5
            },
            'charts_data': self._generate_chart_data(raw_data)
        }
        
        self.logger.debug("Data processing completed")
        return processed_data
    
    def _generate_output_files(self, processed_data: Dict[str, Any],
                              request: ReportRequest,
                              template_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Generate output files in requested formats
        
        Args:
            processed_data: Processed data
            request: Report generation request
            template_data: Template configuration data
            
        Returns:
            Dictionary mapping format to file path
        """
        # Simulate file generation
        time.sleep(1.2)
        
        output_files = {}
        
        # Get output filenames from export settings
        filenames = request.export_settings.get('filenames', {})
        
        for format_name in request.export_settings.get('formats', []):
            if format_name in filenames:
                file_path = filenames[format_name]
                
                # Generate file based on format
                if format_name == 'PDF':
                    self._generate_pdf_report(processed_data, file_path, template_data)
                elif format_name == 'Excel':
                    self._generate_excel_report(processed_data, file_path, template_data)
                elif format_name == 'CSV':
                    self._generate_csv_report(processed_data, file_path, template_data)
                elif format_name == 'JSON':
                    self._generate_json_report(processed_data, file_path, template_data)
                
                output_files[format_name] = file_path
                self.logger.debug(f"Generated {format_name} report: {file_path}")
        
        return output_files
    
    def _generate_pdf_report(self, data: Dict[str, Any], file_path: str, template_data: Dict[str, Any]):
        """Generate PDF report using reportlab"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            # Ensure directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20
            )
            
            # Get styles
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.black
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=20,
                alignment=TA_CENTER,
                textColor=colors.black
            )
            
            section_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                alignment=TA_LEFT,
                textColor=colors.black
            )
            
            # Build content
            story = []
            
            # Header
            story.append(Paragraph("LAPORAN KINERJA KOMPREHENSIF", title_style))
            story.append(Paragraph("PT. PERKEBUNAN NUSANTARA IV", subtitle_style))
            story.append(Spacer(1, 20))
            
            # Summary section
            story.append(Paragraph("RINGKASAN EKSEKUTIF", section_style))
            
            summary_data = data.get('summary', {})
            summary_info = [
                ['Periode Laporan:', data.get('month', 'N/A')],
                ['Total Estate:', str(len(data.get('estates', [])))],
                ['Total Records:', str(summary_data.get('total_records', 0))],
                ['Tanggal Pembuatan:', datetime.now().strftime('%d %B %Y')]
            ]
            
            summary_table = Table(summary_info, colWidths=[2*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # Performance table
            story.append(Paragraph("KINERJA PER ESTATE", section_style))
            
            # Create performance data table
            table_data = [['Kode Estate', 'Nama Estate', 'Target Produksi (Ton)', 'Realisasi Produksi (Ton)', 'Pencapaian (%)', 'Jumlah Karyawan']]
            
            # Add mock data for each estate
            for estate in data.get('estates', []):
                estate_code = estate.get('estate_code', 'N/A')
                estate_name = estate.get('description', 'N/A')
                
                # Generate mock performance data
                target = 1000 + (hash(estate_code) % 500)
                realisasi = target * (0.8 + (hash(estate_code) % 40) / 100)
                pencapaian = (realisasi / target) * 100
                karyawan = 50 + (hash(estate_code) % 30)
                
                table_data.append([
                    estate_code,
                    estate_name,
                    f"{target:,.0f}",
                    f"{realisasi:,.0f}",
                    f"{pencapaian:.1f}%",
                    str(karyawan)
                ])
            
            # Create table
            performance_table = Table(table_data, colWidths=[1*inch, 2*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
            performance_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Estate name left aligned
                ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Numbers right aligned
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            
            story.append(performance_table)
            story.append(Spacer(1, 30))
            
            # Analysis section
            story.append(Paragraph("ANALISIS DETAIL KINERJA", section_style))
            
            analysis_text = f"""
            Berdasarkan data periode {data.get('month', 'N/A')}, berikut adalah analisis kinerja:
            
            1. Pencapaian Target: Rata-rata pencapaian target produksi mencapai 85-95% dari target yang ditetapkan.
            
            2. Produktivitas Karyawan: Produktivitas per karyawan menunjukkan tren yang stabil dengan variasi antar estate.
            
            3. Rekomendasi:
               - Fokus pada peningkatan efisiensi operasional
               - Optimalisasi alokasi sumber daya
               - Peningkatan kapasitas SDM
            """
            
            analysis_para = Paragraph(analysis_text, styles['Normal'])
            story.append(analysis_para)
            
            # Footer info
            story.append(Spacer(1, 30))
            footer_text = f"Laporan dibuat secara otomatis pada {datetime.now().strftime('%d %B %Y, %H:%M:%S')}"
            footer_para = Paragraph(footer_text, styles['Normal'])
            story.append(footer_para)
            
            # Build PDF
            doc.build(story)
            
            self.logger.info(f"PDF report generated successfully: {file_path}")
            
        except ImportError:
            # Fallback to simple text file if reportlab not available
            self.logger.warning("reportlab not available, generating text file instead")
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"LAPORAN KINERJA KOMPREHENSIF\n")
                f.write(f"PT. PERKEBUNAN NUSANTARA IV\n\n")
                f.write(f"Periode: {data.get('month', 'N/A')}\n")
                f.write(f"Total Estate: {len(data.get('estates', []))}\n")
                f.write(f"Tanggal: {datetime.now().strftime('%d %B %Y')}\n\n")
                f.write("KINERJA PER ESTATE:\n")
                for estate in data.get('estates', []):
                    f.write(f"- {estate.get('estate_code', 'N/A')}: {estate.get('description', 'N/A')}\n")
        except Exception as e:
            self.logger.error(f"Error generating PDF report: {e}")
            raise ReportGenerationError(f"Failed to generate PDF report: {e}")
    
    def _generate_excel_report(self, data: Dict[str, Any], file_path: str, template_data: Dict[str, Any]):
        """Generate Excel report using openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Ensure directory exists
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Kinerja"
            
            # Define styles
            title_font = Font(name='Arial', size=16, bold=True)
            subtitle_font = Font(name='Arial', size=12, bold=True)
            header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            data_font = Font(name='Arial', size=10)
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title and header
            ws['A1'] = 'LAPORAN KINERJA KOMPREHENSIF'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            ws.merge_cells('A1:F1')
            
            ws['A2'] = 'PT. PERKEBUNAN NUSANTARA IV'
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = center_align
            ws.merge_cells('A2:F2')
            
            # Summary section
            row = 4
            ws[f'A{row}'] = 'RINGKASAN EKSEKUTIF'
            ws[f'A{row}'].font = subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 2
            summary_data = data.get('summary', {})
            summary_info = [
                ['Periode Laporan:', data.get('month', 'N/A')],
                ['Total Estate:', str(len(data.get('estates', [])))],
                ['Total Records:', str(summary_data.get('total_records', 0))],
                ['Tanggal Pembuatan:', datetime.now().strftime('%d %B %Y')]
            ]
            
            for info in summary_info:
                ws[f'A{row}'] = info[0]
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = info[1]
                row += 1
            
            # Performance table
            row += 2
            ws[f'A{row}'] = 'KINERJA PER ESTATE'
            ws[f'A{row}'].font = subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 2
            # Table headers
            headers = ['Kode Estate', 'Nama Estate', 'Target Produksi (Ton)', 'Realisasi Produksi (Ton)', 'Pencapaian (%)', 'Jumlah Karyawan']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Data rows
            row += 1
            for i, estate in enumerate(data.get('estates', [])):
                estate_code = estate.get('estate_code', 'N/A')
                estate_name = estate.get('description', 'N/A')
                
                # Generate mock performance data
                target = 1000 + (hash(estate_code) % 500)
                realisasi = target * (0.8 + (hash(estate_code) % 40) / 100)
                pencapaian = (realisasi / target) * 100
                karyawan = 50 + (hash(estate_code) % 30)
                
                row_data = [
                    estate_code,
                    estate_name,
                    target,
                    realisasi,
                    f"{pencapaian:.1f}%",
                    karyawan
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # Alignment
                    if col == 1:  # Estate code
                        cell.alignment = center_align
                    elif col == 2:  # Estate name
                        cell.alignment = left_align
                    else:  # Numbers
                        cell.alignment = right_align
                    
                    # Alternating row colors
                    if i % 2 == 1:
                        cell.fill = alt_fill
                
                row += 1
            
            # Analysis section
            row += 2
            ws[f'A{row}'] = 'ANALISIS DETAIL KINERJA'
            ws[f'A{row}'].font = subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 2
            analysis_text = f"""Berdasarkan data periode {data.get('month', 'N/A')}, berikut adalah analisis kinerja:

1. Pencapaian Target: Rata-rata pencapaian target produksi mencapai 85-95% dari target yang ditetapkan.

2. Produktivitas Karyawan: Produktivitas per karyawan menunjukkan tren yang stabil dengan variasi antar estate.

3. Rekomendasi:
   - Fokus pada peningkatan efisiensi operasional
   - Optimalisasi alokasi sumber daya
   - Peningkatan kapasitas SDM"""
            
            ws[f'A{row}'] = analysis_text
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.merge_cells(f'A{row}:F{row + 10}')
            
            # Adjust column widths
            column_widths = [15, 25, 20, 20, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Save workbook
            wb.save(file_path)
            
            self.logger.info(f"Excel report generated successfully: {file_path}")
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            self.logger.warning("openpyxl not available, generating CSV file instead")
            csv_path = file_path.replace('.xlsx', '.csv')
            Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
            
            import csv
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['LAPORAN KINERJA KOMPREHENSIF'])
                writer.writerow(['PT. PERKEBUNAN NUSANTARA IV'])
                writer.writerow([])
                writer.writerow(['Periode Laporan:', data.get('month', 'N/A')])
                writer.writerow(['Total Estate:', str(len(data.get('estates', [])))])
                writer.writerow(['Tanggal:', datetime.now().strftime('%d %B %Y')])
                writer.writerow([])
                writer.writerow(['Kode Estate', 'Nama Estate', 'Target Produksi', 'Realisasi Produksi', 'Pencapaian (%)', 'Jumlah Karyawan'])
                
                for estate in data.get('estates', []):
                    estate_code = estate.get('estate_code', 'N/A')
                    estate_name = estate.get('description', 'N/A')
                    target = 1000 + (hash(estate_code) % 500)
                    realisasi = target * (0.8 + (hash(estate_code) % 40) / 100)
                    pencapaian = (realisasi / target) * 100
                    karyawan = 50 + (hash(estate_code) % 30)
                    
                    writer.writerow([estate_code, estate_name, target, f"{realisasi:.0f}", f"{pencapaian:.1f}%", karyawan])
                    
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            raise ReportGenerationError(f"Failed to generate Excel report: {e}")
    
    def _generate_csv_report(self, data: Dict[str, Any], file_path: str, template_data: Dict[str, Any]):
        """Generate CSV report"""
        # Mock CSV generation
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, 'w') as f:
            f.write("CSV Report\n")
            f.write(f"Generated,{datetime.now()}\n")
            f.write(f"Template,{template_data['template'].name}\n")
    
    def _generate_json_report(self, data: Dict[str, Any], file_path: str, template_data: Dict[str, Any]):
        """Generate JSON report"""
        # Mock JSON generation
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        report_data = {
            'generated_at': datetime.now().isoformat(),
            'template': template_data['template'].name,
            'data': data
        }
        with open(file_path, 'w') as f:
            json.dump(report_data, f, indent=2, default=str)
    
    def _finalize_generation(self, result: ReportResult, output_files: Dict[str, str]):
        """
        Finalize report generation
        
        Args:
            result: Report result object
            output_files: Generated output files
        """
        # Simulate finalization
        time.sleep(0.2)
        
        # Here you would typically:
        # - Clean up temporary files
        # - Update generation logs
        # - Send notifications
        
        self.logger.debug("Report generation finalized")
    
    def _generate_mock_transactions(self, request: ReportRequest) -> List[Dict[str, Any]]:
        """Generate mock transaction data"""
        transactions = []
        for i in range(50):  # Mock 50 transactions
            transactions.append({
                'id': f"TXN{i+1:03d}",
                'estate': request.selected_estates[i % len(request.selected_estates)],
                'date': f"{request.selected_month}-{(i % 28) + 1:02d}",
                'amount': 1000 + (i * 50),
                'type': 'FFB' if i % 2 == 0 else 'CPO'
            })
        return transactions
    
    def _generate_chart_data(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate mock chart data"""
        return {
            'monthly_trend': [100, 120, 110, 130, 125],
            'estate_comparison': {estate: 1000 + (i * 200) for i, estate in enumerate(raw_data['estates'])},
            'transaction_types': {'FFB': 60, 'CPO': 40}
        }
    
    def _update_progress(self, step: int, message: str, detail: str = ""):
        """
        Update progress
        
        Args:
            step: Current step number
            message: Progress message
            detail: Detailed message
        """
        if self.progress_callback:
            try:
                self.progress_callback(step, message, detail)
            except Exception as e:
                self.logger.error(f"Error in progress callback: {e}")
    
    def _check_cancellation(self):
        """Check if cancellation was requested"""
        if self.cancel_requested:
            raise ReportGenerationError("Report generation was cancelled")
    
    def cancel_generation(self):
        """Cancel current report generation"""
        if self.is_generating:
            self.cancel_requested = True
            self.logger.info("Report generation cancellation requested")
    
    def get_generation_status(self) -> Dict[str, Any]:
        """
        Get current generation status
        
        Returns:
            Status information dictionary
        """
        return {
            'is_generating': self.is_generating,
            'cancel_requested': self.cancel_requested,
            'current_template': self.current_request.template.name if self.current_request else None,
            'stats': self.generation_stats.copy()
        }
    
    def cleanup(self):
        """Cleanup service resources"""
        if self.is_generating:
            self.cancel_generation()
        
        # Wait for thread to finish
        if self.generation_thread and self.generation_thread.is_alive():
            self.generation_thread.join(timeout=5.0)
        
        self.logger.info("Report generation service cleaned up")