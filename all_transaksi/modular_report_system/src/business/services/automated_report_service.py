"""
Automated Report Service
Handles automated report generation with predefined parameters
"""

import logging
import schedule
import time
import threading
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Callable, Any
from pathlib import Path

from business.models.report_request import ReportRequest
from business.services.report_generation_service import ReportGenerationService
from business.services.validation_service import ValidationService
from infrastructure.exceptions.custom_exceptions import *


class AutomatedReportService:
    """Service for automated report generation with predefined parameters"""
    
    def __init__(self, report_service: ReportGenerationService, validation_service: ValidationService):
        self.logger = logging.getLogger(__name__)
        self.report_service = report_service
        self.validation_service = validation_service
        self.scheduler_thread = None
        self.is_running = False
        
        # Predefined configurations for automated reports
        self.automated_configs = {
            'unit_9_estate_2b': {
                'name': 'Unit 9 - Estate 2B Monthly Report',
                'description': 'Laporan bulanan otomatis untuk Unit 9 Estate 2B',
                'estates': [{'id': '2B', 'name': 'Estate 2B', 'unit': 9}],
                'template': {'name': 'Monthly Standard Report', 'type': 'monthly'},
                'schedule': 'monthly',  # monthly, weekly, daily
                'schedule_day': 1,  # Day of month for monthly reports
                'export_format': 'excel',
                'include_charts': True,
                'include_summary': True,
                'auto_email': False,
                'email_recipients': [],
                'enabled': True
            }
        }
        
        self.status_callbacks = []
    
    def validate_report_data(self, data, unit=None, estate=None):
        """
        Validate report data for specific unit and estate
        
        Args:
            data: Report data to validate
            unit: Unit number for validation
            estate: Estate identifier for validation
            
        Returns:
            dict: Validation results with status and messages
        """
        try:
            validation_results = {
                'is_valid': True,
                'errors': [],
                'warnings': [],
                'data_quality_score': 100
            }
            
            # Check if data exists
            if not data or len(data) == 0:
                validation_results['is_valid'] = False
                validation_results['errors'].append("Tidak ada data untuk divalidasi")
                validation_results['data_quality_score'] = 0
                return validation_results
            
            # Validate unit and estate specific data
            if unit and estate:
                unit_estate_data = [row for row in data if 
                                  str(row.get('unit', '')).strip() == str(unit).strip() and 
                                  str(row.get('estate', '')).strip() == str(estate).strip()]
                
                if not unit_estate_data:
                    validation_results['warnings'].append(f"Tidak ada data untuk Unit {unit} Estate {estate}")
                    validation_results['data_quality_score'] -= 20
                else:
                    self.logger.info(f"Found {len(unit_estate_data)} records for Unit {unit} Estate {estate}")
            
            # Validate required fields
            required_fields = ['tanggal', 'unit', 'estate', 'transno']
            missing_fields = []
            
            for row in data[:10]:  # Check first 10 rows for performance
                for field in required_fields:
                    if field not in row or not row[field]:
                        if field not in missing_fields:
                            missing_fields.append(field)
            
            if missing_fields:
                validation_results['warnings'].append(f"Field yang hilang: {', '.join(missing_fields)}")
                validation_results['data_quality_score'] -= len(missing_fields) * 10
            
            # Validate data types and formats
            date_errors = 0
            for i, row in enumerate(data[:100]):  # Check first 100 rows
                # Validate date format
                if 'tanggal' in row and row['tanggal']:
                    try:
                        if isinstance(row['tanggal'], str):
                            datetime.strptime(row['tanggal'][:10], '%Y-%m-%d')
                    except ValueError:
                        date_errors += 1
                        if date_errors <= 5:  # Only report first 5 errors
                            validation_results['warnings'].append(f"Format tanggal tidak valid pada baris {i+1}")
            
            if date_errors > 0:
                validation_results['data_quality_score'] -= min(date_errors * 2, 20)
            
            # Check for duplicates
            if len(data) > 1:
                transno_list = [row.get('transno') for row in data if row.get('transno')]
                duplicates = len(transno_list) - len(set(transno_list))
                if duplicates > 0:
                    validation_results['warnings'].append(f"Ditemukan {duplicates} duplikasi transno")
                    validation_results['data_quality_score'] -= min(duplicates, 15)
            
            # Final validation status
            if validation_results['data_quality_score'] < 70:
                validation_results['is_valid'] = False
                validation_results['errors'].append("Kualitas data di bawah standar minimum (70%)")
            
            self.logger.info(f"Data validation completed. Score: {validation_results['data_quality_score']}%")
            return validation_results
            
        except Exception as e:
            self.logger.error(f"Error validating report data: {e}")
            return {
                'is_valid': False,
                'errors': [f"Error validasi: {str(e)}"],
                'warnings': [],
                'data_quality_score': 0
            }
    
    def format_report_output(self, data, format_type='excel', unit=None, estate=None):
        """
        Format report output according to specified format
        
        Args:
            data: Report data to format
            format_type: Output format ('excel', 'pdf', 'csv')
            unit: Unit number for formatting
            estate: Estate identifier for formatting
            
        Returns:
            str: Path to formatted output file
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unit_estate_str = f"Unit{unit}_Estate{estate}" if unit and estate else "AllUnits"
            
            if format_type.lower() == 'excel':
                return self._format_excel_output(data, unit_estate_str, timestamp)
            elif format_type.lower() == 'pdf':
                return self._format_pdf_output(data, unit_estate_str, timestamp)
            elif format_type.lower() == 'csv':
                return self._format_csv_output(data, unit_estate_str, timestamp)
            else:
                raise ValueError(f"Format tidak didukung: {format_type}")
                
        except Exception as e:
            self.logger.error(f"Error formatting report output: {e}")
            raise
    
    def _format_excel_output(self, data, unit_estate_str, timestamp):
        """Format data as Excel file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import os
            
            # Create output directory
            output_dir = os.path.join(os.getcwd(), 'reports', 'automated')
            os.makedirs(output_dir, exist_ok=True)
            
            filename = f"Laporan_{unit_estate_str}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Data"
            
            # Add header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Write headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            self.logger.info(f"Excel report saved: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating Excel output: {e}")
            raise
    
    def _format_csv_output(self, data, unit_estate_str, timestamp):
        """Format data as CSV file"""
        try:
            import pandas as pd
            import os
            
            # Create output directory
            output_dir = os.path.join(os.getcwd(), 'reports', 'automated')
            os.makedirs(output_dir, exist_ok=True)
            
            filename = f"Laporan_{unit_estate_str}_{timestamp}.csv"
            filepath = os.path.join(output_dir, filename)
            
            # Create DataFrame and save as CSV
            df = pd.DataFrame(data)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            self.logger.info(f"CSV report saved: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating CSV output: {e}")
            raise
    
    def _format_pdf_output(self, data, unit_estate_str, timestamp):
        """Format data as PDF file (placeholder implementation)"""
        try:
            import os
            
            # For now, create CSV and note that PDF conversion would need additional libraries
            csv_path = self._format_csv_output(data, unit_estate_str, timestamp)
            
            # Create a simple text-based PDF alternative
            output_dir = os.path.join(os.getcwd(), 'reports', 'automated')
            filename = f"Laporan_{unit_estate_str}_{timestamp}.txt"
            filepath = os.path.join(output_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(f"LAPORAN OTOMATIS - {unit_estate_str}\n")
                f.write(f"Tanggal Generate: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write("="*80 + "\n\n")
                
                if data:
                    # Write headers
                    headers = list(data[0].keys())
                    f.write(" | ".join(headers) + "\n")
                    f.write("-" * (len(" | ".join(headers))) + "\n")
                    
                    # Write data rows
                    for row in data[:100]:  # Limit to first 100 rows for readability
                        values = [str(row.get(header, '')) for header in headers]
                        f.write(" | ".join(values) + "\n")
                    
                    if len(data) > 100:
                        f.write(f"\n... dan {len(data) - 100} baris lainnya\n")
                
                f.write(f"\nTotal Records: {len(data)}\n")
            
            self.logger.info(f"Text report saved: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating PDF output: {e}")
            raise
    
    def add_status_callback(self, callback: Callable[[str, str], None]):
        """Add callback for status updates"""
        self.status_callbacks.append(callback)
    
    def notify_status(self, config_name: str, status: str):
        """Notify all callbacks about status change"""
        for callback in self.status_callbacks:
            try:
                callback(config_name, status)
            except Exception as e:
                self.logger.error(f"Error in status callback: {e}")
    
    def get_automated_configs(self) -> List[Dict]:
        """Get all automated report configurations as a list"""
        return list(self.automated_configs.values())
    
    def update_config(self, config_name: str, updates: Dict):
        """Update automated report configuration"""
        if config_name in self.automated_configs:
            self.automated_configs[config_name].update(updates)
            self.logger.info(f"Updated automated config: {config_name}")
            return True
        return False
    
    def enable_config(self, config_name: str, enabled: bool = True):
        """Enable or disable automated report configuration"""
        if config_name in self.automated_configs:
            self.automated_configs[config_name]['enabled'] = enabled
            status = "enabled" if enabled else "disabled"
            self.logger.info(f"Automated config {config_name} {status}")
            self.notify_status(config_name, f"Configuration {status}")
            return True
        return False
    
    def start_scheduler(self):
        """Start the automated report scheduler"""
        if self.is_running:
            self.logger.warning("Scheduler is already running")
            return
        
        self.is_running = True
        self._setup_schedules()
        
        # Start scheduler in separate thread
        self.scheduler_thread = threading.Thread(target=self._run_scheduler, daemon=True)
        self.scheduler_thread.start()
        
        self.logger.info("Automated report scheduler started")
    
    def stop_scheduler(self):
        """Stop the automated report scheduler"""
        self.is_running = False
        schedule.clear()
        
        if self.scheduler_thread and self.scheduler_thread.is_alive():
            self.scheduler_thread.join(timeout=5)
        
        self.logger.info("Automated report scheduler stopped")
    
    def _setup_schedules(self):
        """Setup schedules for all enabled configurations"""
        schedule.clear()
        
        for config_name, config in self.automated_configs.items():
            if not config.get('enabled', False):
                continue
            
            schedule_type = config.get('schedule', 'monthly')
            
            if schedule_type == 'monthly':
                # Schedule for first day of each month at 8 AM
                schedule.every().month.at("08:00").do(
                    self._generate_automated_report, config_name
                )
            elif schedule_type == 'weekly':
                # Schedule for every Monday at 8 AM
                schedule.every().monday.at("08:00").do(
                    self._generate_automated_report, config_name
                )
            elif schedule_type == 'daily':
                # Schedule for every day at 8 AM
                schedule.every().day.at("08:00").do(
                    self._generate_automated_report, config_name
                )
            
            self.logger.info(f"Scheduled {schedule_type} report for {config_name}")
    
    def _run_scheduler(self):
        """Run the scheduler loop"""
        while self.is_running:
            try:
                schedule.run_pending()
                time.sleep(60)  # Check every minute
            except Exception as e:
                self.logger.error(f"Error in scheduler loop: {e}")
                time.sleep(60)
    
    def generate_report_now(self, config_name: str) -> bool:
        """Generate report immediately for specified configuration"""
        if config_name not in self.automated_configs:
            self.logger.error(f"Configuration not found: {config_name}")
            return False
        
        return self._generate_automated_report(config_name)
    
    def _generate_automated_report(self, config_name: str) -> bool:
        """Generate automated report for specified configuration"""
        try:
            config = self.automated_configs[config_name]
            
            if not config.get('enabled', False):
                self.logger.info(f"Skipping disabled configuration: {config_name}")
                return False
            
            self.logger.info(f"Starting automated report generation: {config_name}")
            self.notify_status(config_name, "Starting report generation...")
            
            # Create report request
            request = self._create_automated_request(config)
            if not request:
                self.notify_status(config_name, "Failed to create report request")
                return False
            
            # Validate request
            validation_result = self.validation_service.validate_report_request(request)
            if not validation_result.is_valid:
                error_msg = f"Validation failed: {', '.join(validation_result.errors)}"
                self.logger.error(error_msg)
                self.notify_status(config_name, f"Validation failed: {error_msg}")
                return False
            
            # Generate report
            self.notify_status(config_name, "Generating report...")
            result = self.report_service.generate_report(request)
            
            if result.success:
                self.logger.info(f"Automated report generated successfully: {config_name}")
                self.notify_status(config_name, f"Report generated successfully: {result.output_path}")
                
                # Send email if configured
                if config.get('auto_email', False) and config.get('email_recipients'):
                    self._send_email_notification(config, result)
                
                return True
            else:
                error_msg = f"Report generation failed: {result.error_message}"
                self.logger.error(error_msg)
                self.notify_status(config_name, error_msg)
                return False
                
        except Exception as e:
            error_msg = f"Error generating automated report {config_name}: {e}"
            self.logger.error(error_msg)
            self.notify_status(config_name, error_msg)
            return False
    
    def _create_automated_request(self, config: Dict) -> Optional[ReportRequest]:
        """Create report request from automated configuration"""
        try:
            # Calculate month (previous month for monthly reports)
            now = datetime.now()
            if config.get('schedule') == 'monthly':
                # For monthly reports, use previous month
                if now.month == 1:
                    month = 12
                    year = now.year - 1
                else:
                    month = now.month - 1
                    year = now.year
            else:
                # For other schedules, use current month
                month = now.month
                year = now.year
            
            # Create output directory with timestamp
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            output_dir = Path("reports") / "automated" / f"{config['name'].replace(' ', '_')}_{timestamp}"
            
            request = ReportRequest(
                month=month,
                year=year,
                estates=config['estates'],
                template=config['template'],
                export_format=config.get('export_format', 'excel'),
                output_directory=str(output_dir),
                include_charts=config.get('include_charts', True),
                include_summary=config.get('include_summary', True)
            )
            
            return request
            
        except Exception as e:
            self.logger.error(f"Error creating automated request: {e}")
            return None
    
    def _send_email_notification(self, config: Dict, result):
        """Send email notification for completed report"""
        try:
            # This would integrate with an email service
            # For now, just log the action
            recipients = config.get('email_recipients', [])
            self.logger.info(f"Would send email notification to {recipients} for report: {result.output_path}")
            
            # TODO: Implement actual email sending
            # - Attach generated report
            # - Include summary information
            # - Send to configured recipients
            
        except Exception as e:
            self.logger.error(f"Error sending email notification: {e}")
    
    def get_next_scheduled_runs(self) -> Dict[str, str]:
        """Get next scheduled run times for all configurations"""
        next_runs = {}
        
        for config_name, config in self.automated_configs.items():
            if not config.get('enabled', False):
                continue
            
            schedule_type = config.get('schedule', 'monthly')
            
            if schedule_type == 'monthly':
                # Calculate next first day of month
                now = datetime.now()
                if now.day == 1 and now.hour < 8:
                    next_run = now.replace(hour=8, minute=0, second=0, microsecond=0)
                else:
                    if now.month == 12:
                        next_run = datetime(now.year + 1, 1, 1, 8, 0, 0)
                    else:
                        next_run = datetime(now.year, now.month + 1, 1, 8, 0, 0)
            elif schedule_type == 'weekly':
                # Calculate next Monday
                now = datetime.now()
                days_ahead = 0 - now.weekday()  # Monday is 0
                if days_ahead <= 0:  # Target day already happened this week
                    days_ahead += 7
                next_run = now + timedelta(days=days_ahead)
                next_run = next_run.replace(hour=8, minute=0, second=0, microsecond=0)
            else:  # daily
                # Calculate next day at 8 AM
                now = datetime.now()
                if now.hour < 8:
                    next_run = now.replace(hour=8, minute=0, second=0, microsecond=0)
                else:
                    next_run = (now + timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0)
            
            next_runs[config_name] = next_run.strftime("%Y-%m-%d %H:%M:%S")
        
        return next_runs
    
    def generate_unit9_estate2b_report(self) -> bool:
        """Generate report for Unit 9 Estate 2B manually"""
        try:
            self.logger.info("Starting manual report generation for Unit 9 Estate 2B")
            
            # Get configuration
            config = self.automated_configs.get('unit_9_estate_2b')
            if not config:
                self.logger.error("Unit 9 Estate 2B configuration not found")
                return False
            
            # Create request
            request = self._create_automated_request(config)
            if not request:
                self.logger.error("Failed to create automated request")
                return False
            
            # Validate request
            validation_result = self.validation_service.validate_report_request(request)
            if not validation_result.is_valid:
                self.logger.error(f"Request validation failed: {validation_result.errors}")
                return False
            
            # Generate report
            self.logger.info("Generating report...")
            result = self.report_service.generate_report(request)
            
            if result.success:
                self.logger.info(f"Report generated successfully: {result.output_path}")
                
                # Notify status callbacks
                for callback in self.status_callbacks:
                    try:
                        callback(f"Report generated: {result.output_path}")
                    except Exception as e:
                        self.logger.error(f"Error in status callback: {e}")
                        
                return result
                
        except Exception as e:
            self.logger.error(f"Error generating Unit 9 Estate 2B report: {e}")
            
            # Notify status callbacks
            for callback in self.status_callbacks:
                try:
                    callback(f"Error: {str(e)}")
                except Exception as callback_error:
                    self.logger.error(f"Error in status callback: {callback_error}")
                    
            raise e
    
    def save_configuration(self, config: Dict[str, Any]) -> Optional[str]:
            """
            Save automated report configuration
            
            Args:
                config: Configuration dictionary
                
            Returns:
                Configuration ID if successful, None otherwise
            """
            try:
                # Generate unique ID for configuration
                config_id = f"config_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                # Add metadata
                config['id'] = config_id
                config['created_at'] = datetime.now().isoformat()
                
                # For now, just log the configuration (in real implementation, save to database/file)
                self.logger.info(f"Saving configuration: {config_id}")
                self.logger.debug(f"Configuration data: {config}")
                
                # Store in memory (temporary implementation)
                if not hasattr(self, '_saved_configs'):
                    self._saved_configs = {}
                
                self._saved_configs[config_id] = config
                
                return config_id
                
            except Exception as e:
                self.logger.error(f"Error saving configuration: {e}")
                return None
        
    def load_configurations(self) -> List[Dict[str, Any]]:
        """
        Load all saved configurations
        
        Returns:
            List of configuration dictionaries
        """
        try:
            if not hasattr(self, '_saved_configs'):
                self._saved_configs = {}
                
            return list(self._saved_configs.values())
            
        except Exception as e:
            self.logger.error(f"Error loading configurations: {e}")
            return []
    
    def delete_configuration(self, config_id: str) -> bool:
        """
        Delete a configuration
        
        Args:
            config_id: Configuration ID to delete
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not hasattr(self, '_saved_configs'):
                self._saved_configs = {}
                
            if config_id in self._saved_configs:
                del self._saved_configs[config_id]
                self.logger.info(f"Configuration deleted: {config_id}")
                return True
            else:
                self.logger.warning(f"Configuration not found: {config_id}")
                return False
                
        except Exception as e:
            self.logger.error(f"Error deleting configuration: {e}")
            return False
    
    def test_configuration(self, config_name: str) -> Dict:
        """Test automated report configuration"""
        try:
            if config_name not in self.automated_configs:
                return {'success': False, 'error': 'Configuration not found'}
            
            config = self.automated_configs[config_name]
            
            # Test request creation
            request = self._create_automated_request(config)
            if not request:
                return {'success': False, 'error': 'Failed to create request'}
            
            # Test validation
            validation_result = self.validation_service.validate_report_request(request)
            if not validation_result.is_valid:
                return {
                    'success': False, 
                    'error': f"Validation failed: {', '.join(validation_result.errors)}"
                }
            
            return {
                'success': True,
                'message': 'Configuration test passed',
                'request_details': {
                    'month': request.month,
                    'year': getattr(request, 'year', datetime.now().year),
                    'estates': [estate['name'] for estate in request.estates],
                    'template': request.template['name'],
                    'export_format': request.export_format
                }
            }
            
        except Exception as e:
            return {'success': False, 'error': f'Test failed: {e}'}